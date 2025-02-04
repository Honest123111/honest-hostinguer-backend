import imaplib
import email
from email.header import decode_header
from datetime import datetime, timedelta
from .models import Load, AddressO, AddressD, Customer, ProcessedEmail, Stop
import requests
import time
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.image import MIMEImage
import re
import openpyxl 
from django.db import transaction
import os 
# Configuration for IMAP and SMTP servers
IMAP_SERVER = "imap.gmail.com"
SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 587
EMAIL_USER = "aventurastorefigures@gmail.com"
EMAIL_PASSWORD = "jnwz asgl hwae mcdi"
amazon_addresses = {}

US_STATES = {
    "AL","AK","AZ","AR","CA","CO","CT","DE","FL","GA","HI","ID","IL","IN","IA",
    "KS","KY","LA","ME","MD","MA","MI","MN","MS","MO","MT","NE","NV","NH","NJ",
    "NM","NY","NC","ND","OH","OK","OR","PA","RI","SC","SD","TN","TX","UT","VT",
    "VA","WA","WV","WI","WY"
}


def send_email(subject, body, recipient):
    try:
        msg = MIMEMultipart()
        msg["From"] = EMAIL_USER
        msg["To"] = recipient
        msg["Subject"] = subject
        msg.attach(MIMEText(body, "plain"))

        server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
        server.starttls()
        server.login(EMAIL_USER, EMAIL_PASSWORD)
        server.sendmail(EMAIL_USER, recipient, msg.as_string())
        server.quit()

        print(f"Email sent to {recipient}")
    except Exception as e:
        print(f"Error sending email: {e}")

# Function to fetch coordinates using Nominatim with Google Maps as backup
def get_coordinates(address, retries=3, delay=5):
    """
    Obtiene las coordenadas y detalles de la dirección usando Nominatim.
    Si no se encuentra, usa Google Maps API como respaldo.
    """
    # Intentar primero con Nominatim
    query = address
    url_nominatim = "https://nominatim.openstreetmap.org/search"
    params_nominatim = {"q": query, "format": "json", "addressdetails": 1, "limit": 1}
    headers = {"User-Agent": "YourAppName/1.0"}

    for attempt in range(retries):
        try:
            response = requests.get(url_nominatim, params=params_nominatim, headers=headers, timeout=10)
            response.raise_for_status()
            data = response.json()

            if data:
                result = data[0]
                return {
                    "coordinates": f"{result['lat']},{result['lon']}",
                    "zip": result.get("address", {}).get("postcode", "Unknown"),
                    "state": result.get("address", {}).get("state", "Unknown"),
                    "city": result.get("address", {}).get("city", "Unknown"),
                }
            else:
                print(f"Nominatim could not find coordinates for the address: {query}")
        except requests.exceptions.RequestException as e:
            print(f"Attempt {attempt + 1} failed to get coordinates with Nominatim for {query}: {e}")
            time.sleep(delay)

    # Si Nominatim falla, usar Google Maps API
    print(f"Nominatim failed for {query}. Switching to Google Maps API.")
    return get_coordinates_from_google(address)


def get_coordinates_from_google(address):
    """
    Obtiene coordenadas usando Google Maps Geocoding API como respaldo.
    """
    google_api_key = "AIzaSyDbt0MU_QRza_GErVNPhbsTL89KL3pAR-w"
    url_google = "https://maps.googleapis.com/maps/api/geocode/json"

    try:
        params = {"address": address, "key": google_api_key}
        response = requests.get(url_google, params=params, timeout=10)
        response.raise_for_status()
        data = response.json()

        if data.get("status") == "OK" and data.get("results"):
            result = data["results"][0]
            location = result["geometry"]["location"]
            address_components = {comp["types"][0]: comp["long_name"] for comp in result["address_components"]}

            return {
                "coordinates": f"{location['lat']},{location['lng']}",
                "zip": address_components.get("postal_code", "Unknown"),
                "state": address_components.get("administrative_area_level_1", "Unknown"),
                "city": address_components.get("locality", "Unknown"),
            }
        else:
            print(f"Google Maps API failed for address {address}: {data.get('status')}")
            return {
                "coordinates": "Unknown",
                "zip": "Unknown",
                "state": "Unknown",
                "city": "Unknown",
            }
    except requests.exceptions.RequestException as e:
        print(f"Google Maps API error: {e}")
        return {
            "coordinates": "Unknown",
            "zip": "Unknown",
            "state": "Unknown",
            "city": "Unknown",
        }


def extract_amazon_truck_codes_from_email(body):
    """
    Extrae códigos de camiones de Amazon desde el cuerpo del correo.
    """
    # Buscar patrones comunes de códigos de Amazon Truck (ejemplo: ONT1, PZxxxx)
    truck_code_pattern = r"\b([A-Z]{2,4}\d{1,6})\b"  # Ejemplo: ONT1, PZ21023
    matches = re.findall(truck_code_pattern, body)
    unique_codes = list(set(matches))  # Eliminar duplicados
    return unique_codes

def update_amazon_addresses_from_email(body):
    """
    Actualiza la lista dinámica de direcciones de Amazon a partir del correo.
    """
    truck_codes = extract_amazon_truck_codes_from_email(body)
    print(f"Extracted truck codes: {truck_codes}")

    for code in truck_codes:
        if code not in amazon_addresses:  # Solo buscar si no está ya almacenado
            print(f"Fetching address for truck code: {code}")
            amazon_addresses[code] = get_coordinates(f"{code} Amazon Sort Center")

# Function to send emails
def send_email(subject, body, recipient):
    try:
        msg = MIMEMultipart("related")  # Permite contenido mixto (texto + imagen)
        msg["From"] = EMAIL_USER
        msg["To"] = recipient
        msg["Subject"] = subject

        # HTML con la imagen al final
        html_body = f"""
        <html>
        <body>
            <p>{body}</p>
            <p><strong>Thank you for choosing our service!</strong></p>
            <p>If you have any questions, feel free to contact us.</p>
            <br>
            <p>Best regards,</p>
            <p><strong>Aventura Technology Team</strong></p>
            <br>
            <img src="cid:footer_image" style="width: 200px; height: auto;">
        </body>
        </html>
        """

        # Adjuntar el contenido HTML
        msg.attach(MIMEText(html_body, "html"))

        # Ruta de la imagen
        image_path = os.path.join("myapp/staticfiles", "HONESTBLACK.png")
        
        if os.path.exists(image_path):
            with open(image_path, "rb") as img_file:
                img = MIMEImage(img_file.read(), _subtype="png")  # Definir tipo de imagen
                img.add_header("Content-ID", "<footer_image>")  # Identificador para HTML
                img.add_header("Content-Disposition", "inline", filename="HONESTBLACK.png")
                msg.attach(img)
        else:
            print(f"⚠️ Image not found at {image_path}. Email sent without an image.")

        # Enviar el correo
        server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
        server.starttls()
        server.login(EMAIL_USER, EMAIL_PASSWORD)
        server.sendmail(EMAIL_USER, recipient, msg.as_string())
        server.quit()

        print(f"✅ Email sent successfully to {recipient}")
    except Exception as e:
        print(f"❌ Error sending email: {e}")
        
        
def fetch_and_create_load():
    """
    Fetch emails and process loads, extrayendo dinámicamente direcciones de Amazon.
    """
    try:
        print("Starting email extraction...")
        mail = imaplib.IMAP4_SSL(IMAP_SERVER)
        mail.login(EMAIL_USER, EMAIL_PASSWORD)
        mail.select("inbox")

        # Define el rango de fechas
        two_days_ago = datetime.now() - timedelta(days=2)
        two_days_ago_str = two_days_ago.strftime("%d-%b-%Y")

        # Buscar correos con los asuntos relevantes
        search_criteria = (
            f'(OR SUBJECT "NEW LOAD" SUBJECT "Amazon Relay Spot Load Capacity Request") '
            f'SINCE "{two_days_ago_str}"'
        )
        status, messages = mail.search(None, search_criteria)
        email_ids = messages[0].split()
        print(f"Found {len(email_ids)} emails matching the criteria.")

        for num in email_ids:
            try:
                print(f"Processing email number {num.decode('utf-8')}...")
                status, msg_data = mail.fetch(num, "(RFC822)")
                for response_part in msg_data:
                    if isinstance(response_part, tuple):
                        msg = email.message_from_bytes(response_part[1])

                        msg_id = msg["Message-ID"]
                        if ProcessedEmail.objects.filter(message_id=msg_id).exists():
                            print(f"Email with ID {msg_id} already processed. Skipping.")
                            continue

                        body = None
                        if msg.is_multipart():
                            for part in msg.walk():
                                if part.get_content_type() == "text/plain":
                                    body = part.get_payload(decode=True)
                                    if body:
                                        body = body.decode('utf-8', 'ignore')
                                        break
                        else:
                            body = msg.get_payload(decode=True)
                            if body:
                                body = body.decode('utf-8', 'ignore')

                        if not body:
                            print(f"Error: Email {num.decode('utf-8')} has no readable body. Skipping.")
                            continue

                        print(f"Email body: {body[:100]}...")

                        # Actualizar la lista de direcciones de Amazon
                        update_amazon_addresses_from_email(body)

                        # Procesar el correo
                        load_data = parse_email_body(body)
                        if load_data:
                            create_load_from_data(load_data)

                        ProcessedEmail.objects.create(message_id=msg_id)
            except Exception as e:
                print(f"Error processing email {num.decode('utf-8')}: {e}")

        mail.logout()
        print("Email extraction completed successfully.")
    except Exception as e:
        print(f"Error in fetch_and_create_load: {e}")

from decimal import Decimal


def parse_email_body(body):
    """
    Extrae los datos principales, incluyendo el honest_id, y la información de los stops del cuerpo
    de un correo que tiene un formato similar al siguiente:

        *Origin Zip*: 15120
        *Origin Address*: XCA2
        *Origin State*: California
        *Destiny Zip*: 13579
        *Destiny Address*: 654 Maple Rd
        *Destiny State*: Florida
        *Customer ID*: 1
        *Equipment Type*: Reefer
        *Loaded Miles*: 1200
        *Total Weight*: 5000
        *Commodity*: Produce
        *Offer*: 2500.00
        *Honest ID*: 47V0621

        *Stop Location*: 101 Apple Ln, Austin, TX
        *Stop Date Time*: 2024-11-24 08:00:00
        *Stop Action Type*: Pickup
        *Stop Estimated Weight*: 1500
        *Stop Quantity*: 10

        (más stops…)
    """
    # Patrón para extraer todas las parejas clave:valor (sin depender de la posición)
    main_pattern = r"\*?(?P<key>[^:*]+)\*?:\s*(?P<value>[^\n]+)"
    matches = re.finditer(main_pattern, body)
    
    data = {}
    for match in matches:
        key = match.group("key").strip().lower().replace(" ", "_")  # ej. "origin_zip"
        value = match.group("value").strip()
        # Si el campo empieza con "stop_" se procesará por separado
        if not key.startswith("stop_"):
            data[key] = value

    # Extraer los stops por separado: se asume que cada stop está definido en un bloque de 5 líneas
    stop_pattern = (
        r"\*Stop Location\*:\s*(?P<location>[^\n]+)\n"
        r"\*Stop Date Time\*:\s*(?P<date_time>[^\n]+)\n"
        r"\*Stop Action Type\*:\s*(?P<action_type>[^\n]+)\n"
        r"\*Stop Estimated Weight\*:\s*(?P<estimated_weight>[^\n]+)\n"
        r"\*Stop Quantity\*:\s*(?P<quantity>[^\n]+)"
    )
    stops = []
    stop_matches = re.finditer(stop_pattern, body, re.MULTILINE)
    for sm in stop_matches:
        try:
            stop = {
                "location": sm.group("location").strip(),
                "date_time": sm.group("date_time").strip(),
                "action_type": sm.group("action_type").strip(),
                "estimated_weight": int(sm.group("estimated_weight").strip()),
                "quantity": int(sm.group("quantity").strip()),
            }
            stops.append(stop)
        except Exception as e:
            print(f"Error processing a stop: {e}")
    data["stops"] = stops

    # Convertir y procesar los campos principales a los tipos deseados
    try:
        processed_data = {
            "origin_zip": data.get("origin_zip", "Unknown"),
            "origin_address": data.get("origin_address", "Unknown"),
            "origin_state": data.get("origin_state", "Unknown"),
            "destiny_zip": data.get("destiny_zip", "Unknown"),
            "destiny_address": data.get("destiny_address", "Unknown"),
            "destiny_state": data.get("destiny_state", "Unknown"),
            "customer_id": int(data.get("customer_id", 0)),
            "equipment_type": data.get("equipment_type", "Unknown"),
            "loaded_miles": int(data.get("loaded_miles", 0)),
            "total_weight": int(data.get("total_weight", 0)),
            "commodity": data.get("commodity", "Unknown"),
            "offer": Decimal(data.get("offer", "0.0")),
            # Agregamos el honest_id (si se encuentra en el correo)
            "honest_id": data.get("honest_id", None),
            "stops": stops,
        }
    except Exception as e:
        print(f"Error processing main fields: {e}")
        return None

    return processed_data


from decimal import Decimal

def create_load_from_data(load_data):
    try:
        # Validar coordenadas para el origen
        origin_coordinates = get_coordinates(load_data.get("origin_address"))
        if origin_coordinates["coordinates"] == "Unknown":
            print(f"Using default values for origin: {load_data.get('origin_address')}")
            origin_coordinates = {"coordinates": "0,0", "zip": "00000", "state": "Unknown", "city": "Unknown"}

        # Validar coordenadas para el destino
        destiny_coordinates = get_coordinates(load_data.get("destiny_address"))
        if destiny_coordinates["coordinates"] == "Unknown":
            print(f"Using default values for destiny: {load_data.get('destiny_address')}")
            destiny_coordinates = {"coordinates": "0,0", "zip": "00000", "state": "Unknown", "city": "Unknown"}

        # Convertir el zip obtenido a un número usando la función auxiliar
        origin_zip_numeric = get_numeric_zip(origin_coordinates["zip"])
        destiny_zip_numeric = get_numeric_zip(destiny_coordinates["zip"])

        # Crear direcciones de origen y destino
        origin = AddressO.objects.create(
            zip_code=origin_zip_numeric,
            address=load_data.get("origin_address"),
            state=origin_coordinates["state"],
            coordinates=origin_coordinates["coordinates"]
        )
        destiny = AddressD.objects.create(
            zip_code=destiny_zip_numeric,
            address=load_data.get("destiny_address"),
            state=destiny_coordinates["state"],
            coordinates=destiny_coordinates["coordinates"]
        )

        # Manejar cliente inexistente
        try:
            customer = Customer.objects.get(id=load_data.get("customer_id"))
        except Customer.DoesNotExist:
            print(f"Customer ID {load_data.get('customer_id')} does not exist. Skipping load creation.")
            return

        # Crear la carga
        load = Load.objects.create(
            origin=origin,
            destiny=destiny,
            equipment_type=load_data.get("equipment_type", "Unknown"),
            customer=customer,
            loaded_miles=load_data.get("loaded_miles", 0),
            total_weight=load_data.get("total_weight", 0),
            commodity=load_data.get("commodity", "Unknown"),
            offer=load_data.get("offer", Decimal("0.0")),
            under_review=True,
            honest_id=load_data.get("honest_id"),
        )
        print(f"Load created successfully: {load.idmmload}")

        # Enviar un email de notificación
        subject = "New Load Created Honest Transportation INC"
        body = f"Se ha creado la carga con ID {load.idmmload} y Honest ID {load_data.get('honest_id')}."
        # Define el destinatario según tu configuración; por ejemplo, el correo del administrador
        recipient = "support@avtechnologyinc.com"
        send_email(subject, body, recipient)
        return load

    except Exception as e:
        print(f"Error creating load: {e}")
        return None





#############################################
# NUEVAS FUNCIONES PARA PROCESAR SPOT LOADS #
#############################################
def process_lane(lane_content):
    """
    Procesa el contenido de la línea Lane y extrae origen, destino y paradas intermedias.
    """
    try:
        if not lane_content or "->" not in lane_content:
            print(f"Error: Formato inválido de Lane. Contenido: {lane_content}")
            return None
        
        lane_content = " ".join(lane_content.split()).replace(",", "").strip()
        lane_parts = [part.strip() for part in lane_content.split("->") if part.strip()]
        
        if len(lane_parts) < 2:
            print(f"Error: Lane no contiene suficientes partes. Contenido: {lane_content}")
            return None
        
        origin = lane_parts[0]
        destination = lane_parts[-1]
        intermediate_stops = lane_parts[1:-1]

        return origin, destination, intermediate_stops
    except Exception as e:
        print(f"Error procesando Lane: {e}")
        return None


def get_location_data(location):
    """
    Fetch location data (e.g., coordinates, zip, state, city) for a given location.
    """
    try:
        # Placeholder para obtener datos del lugar
        data = get_coordinates(location)
        if not data:
            return {
                "coordinates": "Unknown",
                "zip": "Unknown",
                "state": "Unknown",
                "city": "Unknown",
            }
        return data
    except Exception as e:
        print(f"Error fetching location data for {location}: {e}")
        return {
            "coordinates": "Error",
            "zip": "Error",
            "state": "Error",
            "city": "Error",
        }


def parse_spot_load_email_body(body):
    """
    Parse the Spot Load email body to extract load details robustly.
    """
    try:
        # Extraer líneas del cuerpo del correo
        lines = [line.strip() for line in body.splitlines() if line.strip()]
        print("Extracted lines from Spot Load email:")
        for line in lines:
            print(line)

        # Buscar la línea que contiene Lane y reconstruir el contenido
        lane_content = ""
        lane_found = False
        for line in lines:
            if line.lower().startswith("lane:"):
                lane_found = True
                lane_content += line.split(":", 1)[1].strip() if ":" in line else ""
            elif lane_found:
                # Si estamos en la sección de Lane, continuar concatenando
                if line.startswith("Equipment Required:"):
                    break  # Salir cuando llegue al siguiente campo
                lane_content += " " + line.strip()

        # Validar que Lane no esté vacío
        lane_content = lane_content.strip()
        print(f"Raw Lane content: {lane_content}")
        if not lane_content:
            print(f"Error: Formato inválido de Lane. Contenido: {lane_content}")
            return None

        # Procesar el contenido de Lane
        lane_data = process_lane(lane_content)
        if not lane_data:
            return None

        origin, destination, intermediate_stops = lane_data

        # Obtener datos del origen
        origin_data = get_location_data(origin)

        # Obtener datos del destino
        destination_data = get_location_data(destination)

        # Procesar datos de las paradas intermedias
        stops_data = []
        for stop in intermediate_stops:
            stop_data = get_location_data(stop)
            stops_data.append({
                "location": stop,
                "coordinates": stop_data["coordinates"],
                "zip": stop_data["zip"],
                "state": stop_data["state"],
                "city": stop_data["city"],
            })

        # Estructurar los datos finales
        data = {
            "origin_address": origin,
            "origin_data": origin_data,
            "destination_address": destination,
            "destination_data": destination_data,
            "stops": stops_data,
        }
        print(f"Parsed load data: {data}")
        return data
    except Exception as e:
        print(f"Error parsing spot load email: {e}")
        return None




def fetch_and_create_spot_load():
    try:
        print("Iniciando extracción de correos electrónicos Spot Load...")
        mail = imaplib.IMAP4_SSL(IMAP_SERVER)
        mail.login(EMAIL_USER, EMAIL_PASSWORD)
        mail.select("inbox")

        two_days_ago = datetime.now() - timedelta(days=2)
        two_days_ago_str = two_days_ago.strftime("%d-%b-%Y")
        search_criteria = f'(SUBJECT "Amazon Relay Spot Load Capacity Request" SINCE "{two_days_ago_str}")'
        status, messages = mail.search(None, search_criteria)

        for num in messages[0].split():
            status, msg_data = mail.fetch(num, "(RFC822)")
            for response_part in msg_data:
                if isinstance(response_part, tuple):
                    msg = email.message_from_bytes(response_part[1])
                    msg_id = msg["Message-ID"]

                    if ProcessedEmail.objects.filter(message_id=msg_id).exists():
                        print(f"Email con ID {msg_id} ya procesado.")
                        continue

                    body = ""
                    if msg.is_multipart():
                        for part in msg.walk():
                            if part.get_content_type() == "text/plain":
                                body = part.get_payload(decode=True).decode('utf-8', 'ignore')
                                break
                    else:
                        body = msg.get_payload(decode=True).decode('utf-8', 'ignore')

                    load_data = parse_spot_load_email_body(body)
                    if load_data:
                        print(f"Datos del Load: {load_data}")

                    ProcessedEmail.objects.create(message_id=msg_id)

        mail.logout()
        print("Extracción de correos Spot Load completada con éxito.")
    except Exception as e:
        print(f"Error en fetch_and_create_spot_load: {e}")

def fetch_amazon_addresses(truck_codes):
    """
    Llena el diccionario amazon_addresses con direcciones y coordenadas de Amazon.
    """
    url_nominatim = "https://nominatim.openstreetmap.org/search"
    headers = {"User-Agent": "YourApp/1.0"}

    for code in truck_codes:
        if code in amazon_addresses:
            continue  # Ya existe, no buscar de nuevo
        
        query = f"{code} Amazon Sort Center"
        params = {"q": query, "format": "json", "addressdetails": 1, "limit": 1}

        try:
            response = requests.get(url_nominatim, params=params, headers=headers, timeout=10)
            response.raise_for_status()
            data = response.json()

            if data:
                result = data[0]
                amazon_addresses[code] = {
                    "address": result.get("display_name"),
                    "coordinates": f"{result['lat']},{result['lon']}",
                    "zip": result.get("address", {}).get("postcode", "Unknown"),
                    "state": result.get("address", {}).get("state", "Unknown"),
                    "city": result.get("address", {}).get("city", "Unknown"),
                }
                print(f"Address found for {code}: {amazon_addresses[code]}")
            else:
                print(f"No address found for {code}")
                amazon_addresses[code] = {
                    "address": "Unknown",
                    "coordinates": "Unknown",
                    "zip": "Unknown",
                    "state": "Unknown",
                    "city": "Unknown",
                }
            time.sleep(1)  # Respetar los límites de la API
        except requests.exceptions.RequestException as e:
            print(f"Error fetching address for {code}: {e}")
            amazon_addresses[code] = {
                "address": "Error",
                "coordinates": "Error",
                "zip": "Error",
                "state": "Error",
                "city": "Error",
            }

def process_zip_codes(code):
    """
    Formatea y valida los códigos ZIP, incluyendo los prefijos como 'PZ'.
    """
    try:
        # Validar si es un código ZIP válido
        if code.startswith("PZ") and code[2:].isdigit():
            return code[2:]  # Extraer el número después de 'PZ'
        elif code.isdigit():
            return code
        else:
            print(f"Invalid ZIP code format: {code}")
            return "00000"  # Asignar un ZIP predeterminado en caso de fallo
    except Exception as e:
        print(f"Error processing ZIP code {code}: {e}")
        return "00000"
def fetch_zip_code_coordinates(zip_code):
    """
    Busca coordenadas y detalles basados en un código postal en Estados Unidos.
    """
    if not zip_code.isdigit() or len(zip_code) != 5:
        print(f"Invalid ZIP code format: {zip_code}")
        return {
            "coordinates": "Unknown",
            "zip": zip_code,
            "state": "Unknown",
            "city": "Unknown",
        }

    try:
        # Usar Nominatim para buscar por código postal
        url_nominatim = "https://nominatim.openstreetmap.org/search"
        params = {"postalcode": zip_code, "country": "US", "format": "json", "addressdetails": 1, "limit": 1}
        headers = {"User-Agent": "YourAppName/1.0"}
        
        response = requests.get(url_nominatim, params=params, headers=headers, timeout=10)
        response.raise_for_status()
        data = response.json()

        if data:
            result = data[0]
            return {
                "coordinates": f"{result['lat']},{result['lon']}",
                "zip": result.get("address", {}).get("postcode", "Unknown"),
                "state": result.get("address", {}).get("state", "Unknown"),
                "city": result.get("address", {}).get("city", "Unknown"),
            }
        else:
            print(f"No data found for ZIP code: {zip_code}")
    except requests.exceptions.RequestException as e:
        print(f"Error fetching data for ZIP code {zip_code}: {e}")

    return {
        "coordinates": "Unknown",
        "zip": zip_code,
        "state": "Unknown",
        "city": "Unknown",
    }

def clean_address(address):
    """
    Limpia y valida una dirección antes de enviarla a las APIs.
    """
    if not address or not address.strip():
        print("Invalid or empty address provided.")
        return None
    return address.strip()

def clean_truck_code(truck_code):
    """
    Valida códigos de camiones, asegurándose de que sean válidos.
    """
    if not truck_code or not re.match(r"^[A-Z]{2,4}\d+$", truck_code):
        print(f"Invalid truck code: {truck_code}")
        return None
    return truck_code


def fetch_and_create_load_based_on_header():
    """
    - Se conecta al inbox GMail.
    - Busca correos cuyo asunto sea "NEW LOAD", "Amazon Relay Spot Load Capacity Request"
      o "Truck Availability" de los últimos 2 días.
    - Parsean el cuerpo según corresponda y se guarda/actualiza en BD.
    - Se marca cada correo como ProcessedEmail para no repetirlo.
    """
    try:
        print("Iniciando extracción de correos unificada...")

        mail = imaplib.IMAP4_SSL(IMAP_SERVER)
        mail.login(EMAIL_USER, EMAIL_PASSWORD)
        mail.select("inbox")

        # Correos de los últimos 2 días
        two_days_ago = datetime.now() - timedelta(days=2)
        two_days_ago_str = two_days_ago.strftime("%d-%b-%Y")

        # Buscar 3 asuntos en una sola query
        search_criteria = (
            f'(OR (OR SUBJECT "NEW LOAD" SUBJECT "Amazon Relay Spot Load Capacity Request") SUBJECT "Truck Availability") '
            f'SINCE "{two_days_ago_str}"'
        )
        status, messages = mail.search(None, search_criteria)
        if status != "OK":
            print("Error al buscar en la bandeja de entrada.")
            mail.logout()
            return

        email_ids = messages[0].split()
        print(f"Se encontraron {len(email_ids)} correos que coinciden con el criterio.")

        for num in email_ids:
            try:
                status, msg_data = mail.fetch(num, "(RFC822)")
                for response_part in msg_data:
                    if not isinstance(response_part, tuple):
                        continue

                    msg = email.message_from_bytes(response_part[1])
                    msg_id = msg["Message-ID"]
                    subject = msg["Subject"] or ""

                    # Evitar reprocesar
                    if ProcessedEmail.objects.filter(message_id=msg_id).exists():
                        print(f"Correo con ID {msg_id} ya fue procesado. Se omite.")
                        continue

                    # Extraer el cuerpo en texto plano
                    body = None
                    if msg.is_multipart():
                        for part in msg.walk():
                            if part.get_content_type() == "text/plain":
                                body = part.get_payload(decode=True)
                                if body:
                                    body = body.decode('utf-8', 'ignore')
                                    break
                    else:
                        body = msg.get_payload(decode=True)
                        if body:
                            body = body.decode('utf-8', 'ignore')

                    if not body:
                        print("Este correo no contiene cuerpo de texto.")
                        continue

                    print(f"\nProcesando correo con Subject: {subject}")

                    # Decidir según el asunto
                    subj_lower = subject.strip().lower()

                    if subj_lower == "new load":
                        load_data = parse_email_body(body)
                        if load_data:
                            create_load_from_data(load_data)

                    elif "Amazon Relay Spot Load" in subj_lower:
                        spot_data = parse_spot_load_email_body(body)
                        if spot_data:
                            print(f"Spot Load parseado: {spot_data}")
                            #create_load_from_data(load_data)
                            print(spot_data)

                    elif "truck availability" in subj_lower:
                        truck_list = parse_truck_availability_in_multiline(body)
                        if truck_list:
                            # Ejemplo de guardado en tu modelo TruckAvailability
                            for tdata in truck_list:
                                print(f"Camión disponible parseado: {tdata}")
                                # Si tienes un modelo TruckAvailability:
                                # TruckAvailability.objects.create(
                                #     plate_number=tdata["plate_number"],
                                #     equipment=tdata["equipment"],
                                #     availability_date=tdata["availability_date"],
                                #     location=tdata["location"],
                                # )
                            
                            create_loads_for_truck_availability(truck_list)      
                    else:
                        print("Asunto no reconocido. Se omite.")

                    # Marcar el correo como procesado
                    ProcessedEmail.objects.create(message_id=msg_id)

            except Exception as e:
                print(f"Error procesando correo con ID interno {num}: {e}")

        mail.logout()
        print("Extracción de correos unificada finalizada con éxito.")

    except Exception as e:
        print(f"Error en fetch_and_create_load_based_on_header: {e}")


def get_coordinates_from_city_state(city, state):
    """
    Obtiene lat/lon (y opcionalmente zip) a partir de ciudad+estado.
    Usa Nominatim o lo que prefieras. Devuelve un dict con:
      {
        "coordinates": "lat,lon",
        "zip": "12345",
        "state": "CA",
      }
    Ajusta según tu preferencia.
    """
    try:
        query = f"{city}, {state}, USA"
        url = "https://nominatim.openstreetmap.org/search"
        params = {
            "q": query,
            "format": "json",
            "addressdetails": 1,
            "limit": 1
        }
        headers = {"User-Agent": "YourAppName/1.0"}

        response = requests.get(url, params=params, headers=headers, timeout=10)
        response.raise_for_status()
        data = response.json()
        if data:
            result = data[0]
            coords = f"{result['lat']},{result['lon']}"
            # Extraer zip (postcode) si existe
            zip_code = result.get("address", {}).get("postcode", "00000")
            st = result.get("address", {}).get("state", state)
            return {
                "coordinates": coords,
                "zip": zip_code,
                "state": st
            }
        else:
            return {
                "coordinates": "0,0",
                "zip": "00000",
                "state": state
            }
    except Exception as e:
        print(f"Error obteniendo coordenadas: {e}")
        return {
            "coordinates": "0,0",
            "zip": "00000",
            "state": state
        }

def parse_truck_availability_email_body_as_loads(body):
    """
    Parsea el contenido del correo que se titula "Truck Availability"
    pero en realidad contiene datos de cargas.
    Retorna una lista de diccionarios con la info que se necesita para
    crear un Load.
    """
    lines = [line.strip() for line in body.split('\n') if line.strip()]
    loads_data = []

    for line in lines:
        # Aquí en lugar de un split() fijo, llamamos a la función robusta:
        result = parse_line_truck_availability(line)
        if result:
            loads_data.append({
                "honest_id": result["load_id"],               # 47V0621
                "origin_city": result["origin_city"],         # Rancho Cucamonga
                "origin_state": result["origin_state"],       # CA
                "pickup_text": result["pickup"],              # 01/20 13:00 - 13:00 PST
                "destination_city": result["destination_city"], 
                "destination_state": result["destination_state"],
                "delivery_text": result["delivery"],          # 01/23 23:00 - 23:00 EST
            })
        else:
            print(f"Línea ignorada o malformada: {line}")

    return loads_data


def create_loads_for_truck_availability(loads_data):
    """
    Crea un Load por cada diccionario en loads_data, usando:
      - honest_id
      - Ciudades y estados para buscar coordenadas
      - El customer "JB JUNT 3"
      - ...
    """
    from .models import Customer, Load, AddressO, AddressD

    # Obtener (o crear) el customer con nombre "JB JUNT 3" (ajusta a tu DB)
    try:
        jb_junt3_customer = Customer.objects.get(name="JB JUNT 3")
    except Customer.DoesNotExist:
        # Si no existe, crearlo (o manejar de otro modo)
        jb_junt3_customer = Customer.objects.create(
            name="JB JUNT 3",
            email="noreply@jbjunt3.com",
            corporation="JB JUNT 3 Corp",
            phone_number="0000000000"
        )

    # Recorremos cada "carga" parseada
    for item in loads_data:
        honest_id = item.get("honest_id")
        origin_city = item.get("origin_city")
        origin_state = item.get("origin_state")
        destination_city = item.get("destination_city")
        destination_state = item.get("destination_state")

        # Obtener coordenadas para Origen
        origin_coords_data = get_coordinates_from_city_state(origin_city, origin_state)
        origin_zip = origin_coords_data["zip"]
        origin_coords = origin_coords_data["coordinates"]
        # likewise for Destination
        dest_coords_data = get_coordinates_from_city_state(destination_city, destination_state)
        dest_zip = dest_coords_data["zip"]
        dest_coords = dest_coords_data["coordinates"]

        # Crear AddressO
        origin_addr = AddressO.objects.create(
            zip_code=origin_zip if origin_zip.isdigit() else 0,
            address=f"{origin_city}, {origin_state}",
            state=origin_coords_data["state"],
            coordinates=origin_coords,
        )
        # Crear AddressD
        dest_addr = AddressD.objects.create(
            zip_code=dest_zip if dest_zip.isdigit() else 0,
            address=f"{destination_city}, {destination_state}",
            state=dest_coords_data["state"],
            coordinates=dest_coords,
        )

        # Crear la carga
        # No tenemos todos los campos, así que algunos podrían quedar fijos o con valores por defecto
        new_load = Load.objects.create(
            origin=origin_addr,
            destiny=dest_addr,
            equipment_type="Dry Van",  # o algo por defecto si no sabes
            customer=jb_junt3_customer,  # tu "JB JUNT 3"
            loaded_miles=0,             # no lo sabemos, 0?
            total_weight=0,            # lo que sepas
            commodity="Unknown",       # no lo sabemos
            classifications_and_certifications="",
            offer=0.0,
            honest_id=honest_id,
            under_review=True,
        )

        print(f"Creada la carga con honest_id={honest_id} y ID {new_load.idmmload}")

def parse_line_truck_availability(line):
    """
    Parsea una línea tipo:
      47V0621 Rancho Cucamonga CA 01/20 13:00 - 13:00 PST Lansing MI 01/23 23:00 - 23:00 EST

    Devuelve un diccionario con:
      {
        "load_id": "47V0621",
        "origin_city": "Rancho Cucamonga",
        "origin_state": "CA",
        "pickup": "01/20 13:00 - 13:00 PST",
        "destination_city": "Lansing",
        "destination_state": "MI",
        "delivery": "01/23 23:00 - 23:00 EST"
      }

    Ajusta según tu formato real.
    """
    # Lista/Set de estados de EE.UU.
    US_STATES = {
        "AL","AK","AZ","AR","CA","CO","CT","DE","FL","GA","HI","ID","IL","IN","IA",
        "KS","KY","LA","ME","MD","MA","MI","MN","MS","MO","MT","NE","NV","NH","NJ",
        "NM","NY","NC","ND","OH","OK","OR","PA","RI","SC","SD","TN","TX","UT","VT",
        "VA","WA","WV","WI","WY"
    }

    tokens = line.split()
    if not tokens:
        return None

    # 1) El primer token es el "Load ID"
    load_id = tokens[0]

    # 2) Empezamos a recorrer tokens desde 1 en adelante para obtener "Origin City"
    origin_city_tokens = []
    idx = 1
    while idx < len(tokens):
        if tokens[idx] in US_STATES:
            # Llegamos al Origin State
            break
        origin_city_tokens.append(tokens[idx])
        idx += 1
    
    if idx >= len(tokens):
        # No encontramos state para el origin => línea malformada
        return None

    origin_city = " ".join(origin_city_tokens)
    origin_state = tokens[idx]  # Este es el estado
    idx += 1

    # 3) Acumular "Pickup" hasta el siguiente state
    pickup_tokens = []
    while idx < len(tokens):
        if tokens[idx] in US_STATES:
            # Llegamos al Destination State
            break
        pickup_tokens.append(tokens[idx])
        idx += 1

    if idx >= len(tokens):
        # No encontramos destination_state => línea malformada
        return None

    pickup_str = " ".join(pickup_tokens)
    destination_state = tokens[idx]
    idx += 1

    # 4) Acumular lo que viene para la "Delivery"
    #   pero ojo, antes de "delivery" podría estar la Destination City
    #   Normalmente en tu ejemplo: Lansing MI => Lansing es city, MI es state
    #
    # Así que** en realidad** necesitamos la Destination City tokens = ?
    # Fíjate en tu ejemplo: 
    #  ... PST Lansing MI 01/23 ...
    # Lansing es city, MI es su state
    # => en la lógica anterior, hemos tomado "Lansing" como parte de pickup. 
    #
    # Mejor: separemos Destination City igual que origin city: 
    #   - we keep collecting tokens until we see a US_STATES => that is dest city
    #   - then next token is the next US_STATES => oh, wait. We already recognized 
    #     that US_STATES is the "destination_state" => 
    # Con la forma anterior, hemos consumido Lansing como parte de "pickup"? 
    # => Para un caso EXACTO como el screenshot, la "Pickup" NO incluye la city 
    #   sino algo como "01/20 13:00 - 13:00 PST".
    #
    # REFACTOR: lo ideal es:
    #   load_id
    #   origin_city => hasta state
    #   origin_state => next token
    #   pickup => hasta la next city??? 
    #   destination_city => hasta next state
    #   destination_state => next token
    #   delivery => lo que sobra
    #
    # Sin embargo, si tu correo SIEMPRE pone la city y state JUNTOS => 
    # "Pineville NC" => ya detectaste "NC" => la city es 1 token "Pineville"? 
    #
    # Para el ejemplo donde la city "Rancho Cucamonga" => 2 tokens,
    # y "Hopewell Junction" => 2 tokens,
    # la forma robusta es "detectar primer US_STATES => origin_state, 
    #   luego detectamos otra city (uno o varios tokens) => 
    #   luego detectamos 2do US_STATES => destination_state, 
    #   luego lo que sobra es delivery" 
    #
    # Simplificamos asumiendo que "pickup" es la parte que va inmediatamente 
    #   tras origin_state, HASTA que reconozcamos la city de destino, 
    #   que detectaremos al ver un token NO en US_STATES pero que es Capital 
    #   (p.e. Lansing) => esto se complica.
    #
    # => Viene un ejemplo más claro a continuación.
    
    # --   EJEMPLO ALTERNATIVO DE PARSEO PASO A PASO   --
    # 1) load_id = primer token
    # 2) origin_city = tokens hasta state
    # 3) origin_state = ese state
    # 4) pickup = tokens hasta que topemos con "PST", "CST", "MST", "EST" 
    #    (si en tu correo siempre está la zona horaria)
    # 5) A continuación, destination_city = tokens hasta que topemos con US_STATES
    # 6) destination_state = ese US_STATE
    # 7) delivery = lo que quede
    #
    # Basta con que sepas la estructura del correo. 
    # A efectos de DEMO, continuamos con la idea original de 
    # "pickup" es lo que está antes de destination_state, 
    #  y "delivery" es lo que viene después de destination_state.

    # Tomamos lo que quede en tokens como "delivery"
    delivery_tokens = tokens[idx:]
    delivery_str = " ".join(delivery_tokens)

    # Si la city de destino aparece en "pickup_str", 
    # habría que usar un parseo advanced que distinga dónde acaba "pickup" 
    # y empieza la "destination_city". 
    #
    # Para la DEMO, supongamos que tu correo NO incluye la city en esa parte 
    #   (o ya la incluimos en "pickup"?).
    # 
    # Este ejemplo asume:
    #   "pickup_str" = "01/20 13:00 - 13:00 PST Lansing"
    #   "destination_state" = "MI"
    #   "delivery_str" = "01/23 23:00 - 23:00 EST"
    #
    # => entonces "destination_city" = "Lansing" (la última parte de pickup_str).
    # => re-armamos pickup_str sin la última palabra.

    pickup_parts = pickup_str.rsplit(" ", 1)  # separar la última palabra
    if len(pickup_parts) == 2:
        pickup_main, possible_dest_city = pickup_parts
        destination_city = possible_dest_city
        pickup_str = pickup_main
    else:
        # no hay forma de separar => fallback
        destination_city = "Unknown"

    return {
        "load_id": load_id,
        "origin_city": origin_city,
        "origin_state": origin_state,
        "pickup": pickup_str.strip(),
        "destination_city": destination_city,
        "destination_state": destination_state,
        "delivery": delivery_str.strip(),
    }
LOAD_ID_PATTERN = re.compile(r"^(4[78][A-Za-z0-9]{4,6})$")

def parse_truck_availability_in_multiline(body):
    lines = [ln.strip() for ln in body.splitlines() if ln.strip()]

    loads_data = []
    i = 0
    while i < len(lines):
        line = lines[i]
        # Detectar si es un load_id tipo "47K4911" o "48C1600"
        if LOAD_ID_PATTERN.match(line):
            # Verificamos que haya al menos 7 líneas más
            if i+6 < len(lines):
                chunk = lines[i:i+7]
                i += 7  # avanzamos el cursor
                data = {
                    "honest_id":         chunk[0],  # load_id
                    "origin_city":       chunk[1],
                    "origin_state":      chunk[2],
                    "pickup":            chunk[3],
                    "destination_city":  chunk[4],
                    "destination_state": chunk[5],
                    "delivery":          chunk[6],
                }
                loads_data.append(data)
            else:
                # Quedan menos de 7 líneas, se interrumpe
                break
        else:
            # No es un load_id, avanzamos
            i += 1

    return loads_data

def read_new_load_excel(file_path):
    """
    Lee un Excel con formato de 'NEW LOAD' y crea los registros en BD.
    Suponiendo que la primera fila es de cabecera y que las columnas 
    son (LoadID, OriginZip, OriginAddr, OriginState, DestinyZip, ...).
    """
    # Abrir el workbook
    wb = openpyxl.load_workbook(file_path)
    # Asumimos que la hoja activa (active) o con un nombre fijo
    ws = wb.active

    # Empezamos a leer desde la fila 2 (asumiendo fila 1 = encabezado)
    with transaction.atomic():
        for row_idx in range(2, ws.max_row + 1):
            load_id          = ws.cell(row=row_idx, column=1).value
            origin_zip       = ws.cell(row=row_idx, column=2).value
            origin_address   = ws.cell(row=row_idx, column=3).value
            origin_state     = ws.cell(row=row_idx, column=4).value
            destiny_zip      = ws.cell(row=row_idx, column=5).value
            destiny_address  = ws.cell(row=row_idx, column=6).value
            destiny_state    = ws.cell(row=row_idx, column=7).value
            customer_id      = ws.cell(row=row_idx, column=8).value
            equipment_type   = ws.cell(row=row_idx, column=9).value
            loaded_miles     = ws.cell(row=row_idx, column=10).value
            total_weight     = ws.cell(row=row_idx, column=11).value
            commodity        = ws.cell(row=row_idx, column=12).value
            offer            = ws.cell(row=row_idx, column=13).value

            # Convertir a tipos apropiados si es necesario:
            if loaded_miles is None: loaded_miles = 0
            if total_weight is None: total_weight = 0
            if offer is None: offer = 0.0

            # Manejar la creación en BD
            try:
                # Buscamos o creamos direcciones
                origin_obj = AddressO.objects.create(
                    zip_code=origin_zip if origin_zip else "00000",
                    address=origin_address or "Unknown",
                    state=origin_state or "Unknown",
                    coordinates="0,0"
                )
                destiny_obj = AddressD.objects.create(
                    zip_code=destiny_zip if destiny_zip else "00000",
                    address=destiny_address or "Unknown",
                    state=destiny_state or "Unknown",
                    coordinates="0,0"
                )

                # Obtener el customer
                customer = None
                if customer_id:
                    customer = Customer.objects.filter(id=customer_id).first()

                load = Load.objects.create(
                    origin=origin_obj,
                    destiny=destiny_obj,
                    equipment_type=equipment_type or "Dry Van",
                    customer=customer,
                    loaded_miles=loaded_miles,
                    total_weight=total_weight,
                    commodity=commodity or "Unknown",
                    offer=offer,
                    under_review=True,
                    # honest_id=load_id, (si tienes el campo)
                )
                print(f"[NEW LOAD] Creado Load ID {load.idmmload} con {load_id}")
            except Exception as e:
                print(f"Error creando Load en fila {row_idx}: {e}")


def read_spot_load_excel(file_path):
    """
    Lee un Excel con formato 'Amazon Relay Spot Load Capacity Request'.
    Ajustar columnas según tu formato real.
    """
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    try:
        amazoncustomer = Customer.objects.get(name="Amazon US")
    except Customer.DoesNotExist:
        # Si no existe, crearlo (o manejar de otro modo)
        amazoncustomer = Customer.objects.create(
            name="Amazon US",
            email="amazon@amazon.com",
            corporation="Amazon US",
            phone_number="0000000000"
        )
    with transaction.atomic():
        for row_idx in range(2, ws.max_row + 1):
            # Ejemplo de columnas (Lane, Equipment, Weight, etc.)
            lane       = ws.cell(row=row_idx, column=1).value
            originzip  = ws.cell(row=row_idx, column=2).value
            originaddress  = ws.cell(row=row_idx, column=3).value
            originstate  = ws.cell(row=row_idx, column=4).value
            detinyzip  = ws.cell(row=row_idx, column=5).value
            destinyaddress  = ws.cell(row=row_idx, column=6).value
            destinystate  = ws.cell(row=row_idx, column=7).value
            equipment     = ws.cell(row=row_idx, column=8).value
            weight  = ws.cell(row=row_idx, column=9).value
            # etc...

            # Parsear 'lane' si incluye orígenes/destinos
            # y crear direcciones, stops, etc.
            # Este es un ejemplo básico:
            if lane:
                parts = lane.split("->")
                origin_str = parts[0].strip() if len(parts) > 0 else "Unknown"
                destiny_str= parts[-1].strip() if len(parts) > 1 else "Unknown"
                origincoordinates = get_coordinates_from_city_state(originaddress,originzip)
                destinycoordinates = get_coordinates_from_city_state(originaddress,originzip)
                # Lógica para crear addresses...
                origin_obj = AddressO.objects.create(
                    zip_code=originzip, address=origin_str, state=originstate, coordinates=origincoordinates["coordinates"]
                )
                destiny_obj= AddressD.objects.create(
                    zip_code=detinyzip, address=destiny_str, state=destinystate, coordinates=destinycoordinates["coordinates"]
                )

                load = Load.objects.create(
                    origin=origin_obj,
                    destiny=destiny_obj,
                    equipment_type=equipment or "Dry Van",
                    total_weight=weight or 0,
                    under_review=True,
                    customer=amazoncustomer,
                )
                print(f"[SPOT LOAD] Creado Load ID {load.idmmload} con lane={lane}")


def read_truck_availability_excel(file_path):
    """
    Lee un Excel con formato 'Truck Availability'.
    Asumimos que cada fila corresponde a un Load distinto
    (o la lógica que necesites).
    """
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    from .models import Customer, Load, AddressO, AddressD

    # Obtener (o crear) el customer con nombre "JB JUNT 3" (ajusta a tu DB)
    try:
        jb_junt3_customer = Customer.objects.get(name="JB JUNT 3")
    except Customer.DoesNotExist:
        # Si no existe, crearlo (o manejar de otro modo)
        jb_junt3_customer = Customer.objects.create(
            name="JB JUNT 3",
            email="noreply@jbjunt3.com",
            corporation="JB JUNT 3 Corp",
            phone_number="0000000000"
        )
    with transaction.atomic():
        for row_idx in range(2, ws.max_row + 1):
            # Ejemplo: (honest_id, origin_city, origin_state, pickup, 
            #           destination_city, destination_state, delivery)
            honest_id           = ws.cell(row=row_idx, column=1).value
            origin_city         = ws.cell(row=row_idx, column=2).value
            origin_state        = ws.cell(row=row_idx, column=3).value
            pickup              = ws.cell(row=row_idx, column=4).value
            destination_city    = ws.cell(row=row_idx, column=5).value
            destination_state   = ws.cell(row=row_idx, column=6).value
            delivery            = ws.cell(row=row_idx, column=7).value
            destinyC=get_coordinates_from_city_state(origin_city,origin_state)
            originC=get_coordinates_from_city_state(destination_city,destination_state)
            # Crear addresses
            origin_obj = AddressO.objects.create(
                zip_code="00000",
                address=f"{origin_city}, {origin_state}",
                state=origin_state or "Unknown",
                coordinates=originC["coordinates"]
            )
            destiny_obj = AddressD.objects.create(
                zip_code="00000",
                address=f"{destination_city}, {destination_state}",
                state=destination_state or "Unknown",
                coordinates=destinyC["coordinates"]
            )

            load = Load.objects.create(
                origin=origin_obj,
                destiny=destiny_obj,
                equipment_type="Dry Van",
                # ...
                honest_id=honest_id,
                under_review=True,
                customer=jb_junt3_customer,
            )
            print(f"[TRUCK AVAILABILITY] Creado Load ID {load.idmmload} con honest_id={honest_id}")

def get_numeric_zip(zip_str):
    """
    Dado un código postal en forma de cadena, extrae solo los dígitos y lo convierte a entero.
    Si no se encuentran dígitos, devuelve 0.
    """
    if zip_str is None:
        return 0
    digits = ''.join(filter(str.isdigit, zip_str))
    return int(digits) if digits else 0
